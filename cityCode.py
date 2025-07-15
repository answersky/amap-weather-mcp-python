from openpyxl import load_workbook

# 加载Excel文件
wb = load_workbook('AMap_adcode_citycode.xlsx')

# 获取活动工作表或指定工作表
ws = wb.active  # 或者 wb['Sheet1']

def getCityCode(cityName):
    # 遍历所有行和列
    for row in ws.iter_rows():
        name = ''
        code = ''
        for cell in row:
            if cell.column == 1:
                name = cell.value
            if cell.column == 2:
                code = cell.value

        if cityName in name:
            return code

# print(getCityCode('武汉市'))